import openpyxl
"""
https://openpyxl.readthedocs.io/en/stable/
https://www.pythonexcel.com/openpyxl.php
"""

o = openpyxl.load_workbook("../testData/TestData.xlsx")

# o.create_sheet("Search",2)
# vk  = o.sheetnames
# print(vk)
# o.save("../testData/TestData.xlsx")



sh1=o["Login"]
sh1['B2']="www.amazon.in"
o.save("../testData/TestData.xlsx")


#
sh=o["Login"]

val =sh['A2'].value
print(val)
#
#
#
rows = sh.max_row
columns = sh.max_column

print(rows)
print(columns)




#
# # object of the sheet
# sh =o["Login"]
# print(sh.title)



